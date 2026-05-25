"""Export/import horario data as Excel (shared format) or PDF."""

from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import Asignatura, Profesor, Sesion, Titulacion

EXPORT_VERSION = 1
DIAS_ORDER = ['LUN', 'MAR', 'MIE', 'JUE', 'VIE']
DIAS_FULL = {
    'LUN': 'LUNES',
    'MAR': 'MARTES',
    'MIE': 'MIERCOLES',
    'JUE': 'JUEVES',
    'VIE': 'VIERNES',
}
SHEET_INFO = '_info'
SHEET_TITULACIONES = 'Titulaciones'
SHEET_PROFESORES = 'Profesores'
SHEET_SESIONES = 'Sesiones'
SHEET_ASIGNATURAS = 'Asignaturas'
SHEET_ASIG_SESIONES = 'Asignatura_Sesiones'

DIAS_LABEL = dict(Sesion.DIAS_SEMANA)

CURSO_FILLS = {
    1: 'FFF2CC',
    2: 'D9E2F3',
    3: 'FCE4D6',
    4: 'FFE699',
    5: 'E2EFDA',
}

THIN_BORDER = Border(
    left=Side(style='thin', color='999999'),
    right=Side(style='thin', color='999999'),
    top=Side(style='thin', color='999999'),
    bottom=Side(style='thin', color='999999'),
)
HEADER_FILL = PatternFill('solid', fgColor='7F2020')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(bold=True, size=13)
GROUP_FONT = Font(bold=True, size=9)
CELL_FONT = Font(size=9)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _bool_str(value):
    return '1' if value else '0'


def _parse_bool(value):
    return str(value).strip().lower() in ('1', 'true', 'si', 'sí', 'yes')


def _time_str(t):
    return t.strftime('%H:%M') if t else ''


def _slot_label(hora_inicio):
    h, m = hora_inicio.hour, hora_inicio.minute
    return f'{h}:{m:02d}' if m else f'{h}:00'


def _prof_short(profesor):
    if not profesor:
        return ''
    base = profesor.apellidos.split(',')[0].split()[0]
    return base.replace(' ', '').upper()


def _grup_label(tit_codigo, curso):
    return f'{curso}{tit_codigo.upper()}'


def _collect_merged_schedule(user, cuatrimestre):
    asignaturas = (
        Asignatura.objects.filter(creado_por=user, cuatrimestre=cuatrimestre)
        .select_related('titulacion', 'profesor')
        .prefetch_related('sesiones')
    )

    slots = {}
    groups = {}
    cells = {}

    for asig in asignaturas:
        gkey = (asig.titulacion_id, asig.curso)
        groups[gkey] = (asig.titulacion, asig.curso)
        for ses in asig.sesiones.all():
            skey = (ses.hora_inicio, ses.hora_fin)
            slots[skey] = skey
            ckey = (skey, asig.titulacion_id, asig.curso, ses.dia)
            text = f'{asig.nombre}\n{_prof_short(asig.profesor)}'
            cells.setdefault(ckey, []).append(text)

    sorted_slots = sorted(slots.keys())
    sorted_groups = sorted(
        groups.values(),
        key=lambda x: (x[1], x[0].codigo),
    )
    return sorted_slots, sorted_groups, cells


def _style_cell(cell, fill_hex=None, bold=False, header=False):
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    if header:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    elif fill_hex:
        cell.fill = PatternFill('solid', fgColor=fill_hex)
        cell.font = GROUP_FONT if bold else CELL_FONT
    else:
        cell.font = GROUP_FONT if bold else CELL_FONT


def _write_merged_horario_sheet(ws, user, cuatrimestre_num, cuat_label):
    year = datetime.now().year
    slots, groups, cells = _collect_merged_schedule(user, cuatrimestre_num)
    if not slots or not groups:
        return False

    last_col = 2 + len(DIAS_ORDER)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title = ws.cell(
        row=1, column=1,
        value=f'HORARIOS {year}-{year + 1} {cuat_label.upper()}',
    )
    title.font = TITLE_FONT
    title.alignment = CENTER

    header_row = 2
    headers = ['HORA', 'GRUP'] + [DIAS_FULL[d] for d in DIAS_ORDER]
    for idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=label)
        _style_cell(cell, header=True)

    row = header_row + 1
    for slot in slots:
        slot_start = row
        hora_label = _slot_label(slot[0])
        for tit, curso in groups:
            ws.cell(row=row, column=2, value=_grup_label(tit.codigo, curso))
            _style_cell(ws.cell(row=row, column=2), fill_hex=CURSO_FILLS.get(curso), bold=True)
            for d_idx, dia in enumerate(DIAS_ORDER):
                ckey = (slot, tit.id, curso, dia)
                content = '\n\n'.join(cells.get(ckey, []))
                cell = ws.cell(row=row, column=3 + d_idx, value=content or None)
                fill = CURSO_FILLS.get(curso) if content else None
                _style_cell(cell, fill_hex=fill)
            row += 1
        if row - 1 >= slot_start:
            ws.merge_cells(start_row=slot_start, start_column=1, end_row=row - 1, end_column=1)
            hora_cell = ws.cell(row=slot_start, column=1, value=hora_label)
            _style_cell(hora_cell, bold=True)
            hora_cell.alignment = CENTER

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    for i in range(len(DIAS_ORDER)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 22
    ws.freeze_panes = 'C3'
    return True


def _append_data_sheets(wb, user):
    ws_info = wb.create_sheet(SHEET_INFO)
    ws_info.append(['clave', 'valor'])
    ws_info.append(['version', EXPORT_VERSION])
    ws_info.append(['exportado', timezone.now().isoformat()])
    ws_info.append(['usuario', user.username])

    ws_tit = wb.create_sheet(SHEET_TITULACIONES)
    ws_tit.append(['codigo', 'nombre'])
    for t in Titulacion.objects.filter(creado_por=user).order_by('nombre'):
        ws_tit.append([t.codigo, t.nombre])

    ws_prof = wb.create_sheet(SHEET_PROFESORES)
    ws_prof.append(['nombre', 'apellidos', 'email'])
    for p in Profesor.objects.filter(creado_por=user).order_by('apellidos', 'nombre'):
        ws_prof.append([p.nombre, p.apellidos, p.email or ''])

    ws_ses = wb.create_sheet(SHEET_SESIONES)
    ws_ses.append(['dia', 'hora_inicio', 'hora_fin'])
    for s in Sesion.objects.filter(creado_por=user).order_by('dia', 'hora_inicio'):
        ws_ses.append([s.dia, _time_str(s.hora_inicio), _time_str(s.hora_fin)])

    ws_asig = wb.create_sheet(SHEET_ASIGNATURAS)
    ws_asig.append([
        'codigo', 'nombre', 'titulacion_codigo', 'curso', 'cuatrimestre',
        'es_electiva', 'profesor_nombre', 'profesor_apellidos',
    ])
    asignaturas = (
        Asignatura.objects.filter(creado_por=user)
        .select_related('titulacion', 'profesor')
        .order_by('codigo')
    )
    for a in asignaturas:
        ws_asig.append([
            a.codigo,
            a.nombre,
            a.titulacion.codigo,
            a.curso,
            a.cuatrimestre,
            _bool_str(a.es_electiva),
            a.profesor.nombre if a.profesor else '',
            a.profesor.apellidos if a.profesor else '',
        ])

    ws_link = wb.create_sheet(SHEET_ASIG_SESIONES)
    ws_link.append(['asignatura_codigo', 'dia', 'hora_inicio', 'hora_fin'])
    for a in asignaturas.prefetch_related('sesiones'):
        for s in a.sesiones.all():
            ws_link.append([a.codigo, s.dia, _time_str(s.hora_inicio), _time_str(s.hora_fin)])


def export_to_excel(user):
    wb = Workbook()
    wb.remove(wb.active)

    sheet_index = 0
    for cuat_num, cuat_label in Asignatura.CUATRIMESTRES:
        ws = wb.create_sheet(f'Horario_{cuat_num}', sheet_index)
        if _write_merged_horario_sheet(ws, user, cuat_num, cuat_label):
            sheet_index += 1
        else:
            wb.remove(ws)

    if sheet_index == 0:
        ws = wb.create_sheet('Horario', 0)
        ws.cell(row=1, column=1, value='Sin horarios configurados para exportar.')

    _append_data_sheets(wb, user)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def excel_response(user):
    buffer = export_to_excel(user)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="horario_{user.username}_{datetime.now():%Y%m%d}.xlsx"'
    )
    return response


def export_to_pdf(user):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20)
    styles = getSampleStyleSheet()
    story = []
    year = datetime.now().year
    has_content = False

    for cuat_num, cuat_label in Asignatura.CUATRIMESTRES:
        slots, groups, cells = _collect_merged_schedule(user, cuat_num)
        if not slots or not groups:
            continue
        has_content = True
        story.append(Paragraph(
            f'HORARIOS {year}-{year + 1} {cuat_label.upper()}',
            styles['Title'],
        ))
        story.append(Spacer(1, 8))
        header = ['HORA', 'GRUP'] + [DIAS_FULL[d] for d in DIAS_ORDER]
        rows = [header]
        for slot in slots:
            first = True
            for tit, curso in groups:
                row = [
                    _slot_label(slot[0]) if first else '',
                    _grup_label(tit.codigo, curso),
                ]
                first = False
                for dia in DIAS_ORDER:
                    ckey = (slot, tit.id, curso, dia)
                    row.append('\n\n'.join(cells.get(ckey, [])) or '')
                rows.append(row)

        table = Table(rows, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7F2020')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        story.append(table)
        story.append(Spacer(1, 16))

    if not has_content:
        story.append(Paragraph('Sin horarios configurados.', styles['Normal']))

    doc.build(story)
    buffer.seek(0)
    return buffer


def pdf_response(user):
    buffer = export_to_pdf(user)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="horario_{user.username}_{datetime.now():%Y%m%d}.pdf"'
    )
    return response


def _rows(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return []
    headers = [str(h).strip().lower() if h is not None else '' for h in data[0]]
    rows = []
    for raw in data[1:]:
        if not any(raw):
            continue
        row = {}
        for i, key in enumerate(headers):
            if key:
                row[key] = raw[i] if i < len(raw) else ''
        rows.append(row)
    return rows


def _parse_time(value):
    if value is None or value == '':
        return None
    if hasattr(value, 'hour'):
        return value
    text = str(value).strip()
    for fmt in ('%H:%M:%S', '%H:%M'):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    raise ValueError(f'Hora invalida: {value}')


def import_from_excel(user, uploaded_file):
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    skipped = []
    created = {'titulaciones': 0, 'profesores': 0, 'sesiones': 0, 'asignaturas': 0, 'enlaces': 0}

    tit_map = {}
    for row in _rows(wb, SHEET_TITULACIONES):
        codigo = str(row.get('codigo', '')).strip()
        nombre = str(row.get('nombre', '')).strip()
        if not codigo or not nombre:
            continue
        existing = Titulacion.objects.filter(codigo=codigo, creado_por=user).first()
        if existing:
            skipped.append(f'Titulacion "{codigo}" ya existe; se reutiliza.')
            tit_map[codigo] = existing
        else:
            tit = Titulacion.objects.create(codigo=codigo, nombre=nombre, creado_por=user)
            tit_map[codigo] = tit
            created['titulaciones'] += 1

    prof_map = {}
    for p in Profesor.objects.filter(creado_por=user):
        prof_map[(p.nombre.strip().lower(), p.apellidos.strip().lower())] = p

    for row in _rows(wb, SHEET_PROFESORES):
        nombre = str(row.get('nombre', '')).strip()
        apellidos = str(row.get('apellidos', '')).strip()
        email = str(row.get('email', '')).strip() or None
        if not nombre or not apellidos:
            continue
        key = (nombre.lower(), apellidos.lower())
        if key in prof_map:
            skipped.append(f'Profesor "{apellidos}, {nombre}" ya existe; se reutiliza.')
        else:
            prof = Profesor.objects.create(
                nombre=nombre, apellidos=apellidos, email=email, creado_por=user,
            )
            prof_map[key] = prof
            created['profesores'] += 1

    ses_map = {}
    for s in Sesion.objects.filter(creado_por=user):
        ses_map[(s.dia, s.hora_inicio, s.hora_fin)] = s

    for row in _rows(wb, SHEET_SESIONES):
        dia = str(row.get('dia', '')).strip().upper()
        if dia not in DIAS_ORDER:
            skipped.append(f'Sesion ignorada: dia invalido "{dia}".')
            continue
        try:
            hi = _parse_time(row.get('hora_inicio'))
            hf = _parse_time(row.get('hora_fin'))
        except ValueError as exc:
            skipped.append(str(exc))
            continue
        key = (dia, hi, hf)
        if key in ses_map:
            skipped.append(
                f'Sesion {DIAS_LABEL.get(dia, dia)} {_time_str(hi)}-{_time_str(hf)} ya existe; se reutiliza.'
            )
        else:
            ses = Sesion.objects.create(dia=dia, hora_inicio=hi, hora_fin=hf, creado_por=user)
            ses_map[key] = ses
            created['sesiones'] += 1

    asig_map = {}
    for a in Asignatura.objects.filter(creado_por=user):
        asig_map[a.codigo] = a

    for row in _rows(wb, SHEET_ASIGNATURAS):
        codigo = str(row.get('codigo', '')).strip()
        if not codigo:
            continue
        if codigo in asig_map:
            skipped.append(f'Asignatura "{codigo}" ya existe; se omite.')
            continue
        tit_codigo = str(row.get('titulacion_codigo', '')).strip()
        tit = tit_map.get(tit_codigo) or Titulacion.objects.filter(
            codigo=tit_codigo, creado_por=user,
        ).first()
        if not tit:
            skipped.append(f'Asignatura "{codigo}" omitida: titulacion "{tit_codigo}" no encontrada.')
            continue
        prof = None
        pn = str(row.get('profesor_nombre', '')).strip()
        pa = str(row.get('profesor_apellidos', '')).strip()
        if pn and pa:
            prof = prof_map.get((pn.lower(), pa.lower()))
        try:
            curso = int(row.get('curso', 1))
            cuatrimestre = int(row.get('cuatrimestre', 1))
        except (TypeError, ValueError):
            skipped.append(f'Asignatura "{codigo}" omitida: curso/cuatrimestre invalido.')
            continue
        asig = Asignatura.objects.create(
            codigo=codigo,
            nombre=str(row.get('nombre', codigo)).strip(),
            titulacion=tit,
            curso=curso,
            cuatrimestre=cuatrimestre,
            es_electiva=_parse_bool(row.get('es_electiva', '0')),
            profesor=prof,
            creado_por=user,
        )
        asig_map[codigo] = asig
        created['asignaturas'] += 1

    for row in _rows(wb, SHEET_ASIG_SESIONES):
        codigo = str(row.get('asignatura_codigo', '')).strip()
        asig = asig_map.get(codigo)
        if not asig:
            continue
        dia = str(row.get('dia', '')).strip().upper()
        try:
            hi = _parse_time(row.get('hora_inicio'))
            hf = _parse_time(row.get('hora_fin'))
        except ValueError:
            continue
        ses = ses_map.get((dia, hi, hf))
        if ses and not asig.sesiones.filter(pk=ses.pk).exists():
            asig.sesiones.add(ses)
            created['enlaces'] += 1

    return {'created': created, 'skipped': skipped}
